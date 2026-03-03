# -*- coding: utf-8 -*-
"""
OMR 이미지들을 읽어 채점하고, 엑셀에서 해당 사번에 점수 반영.
- config.yaml 에 정답 키·엑셀 컬럼명 등 설정.
- 이미지에서 QR로 사번을 읽거나, 파일명에서 사번 추출.
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import yaml
except ImportError:
    yaml = None

from read_omr import read_omr, get_id_from_filename


def load_config(config_path: str | Path) -> dict:
    path = Path(config_path)
    if not path.exists():
        raise FileNotFoundError(f"설정 파일이 없습니다: {config_path}")
    with open(path, "r", encoding="utf-8") as f:
        if yaml is None:
            raise ImportError("PyYAML이 필요합니다: pip install pyyaml")
        return yaml.safe_load(f)


def score_answers(answers: list[int], answer_key: list[int], points_per: float = 1.0) -> float:
    """문항별 정답과 비교해 총점 계산."""
    total = 0.0
    n = min(len(answers), len(answer_key))
    for i in range(n):
        if answers[i] == answer_key[i]:
            total += points_per
    return total


def load_answer_key_from_excel(
    excel_path: str | Path,
    exam_code: str,
    num_questions: int,
) -> list[int]:
    """
    엑셀 시트1의 1열(시험번호), 2열(시험정답)에서 해당 시험번호의 정답을 읽어 리스트로 반환.

    예시 구조 (시트1):
        A열: 시험번호
        B열: 시험정답  (예: 12341234... 또는 '1 2 3 4 ...')
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]  # 시트1

    target = str(exam_code).strip()
    match_value: str | None = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 2:
            continue
        code_cell, key_cell = row[0], row[1]
        if code_cell is None or key_cell is None:
            continue
        if str(code_cell).strip() == target:
            match_value = str(key_cell)
            break

    if match_value is None:
        raise ValueError(f"시트1에서 시험번호 '{exam_code}' 를 찾을 수 없습니다.")

    import re

    digits = re.findall(r"[1-9]", match_value)
    if len(digits) < num_questions:
        raise ValueError(
            f\"시험번호 '{exam_code}' 의 시험정답에서 {num_questions}개 이상의 선택지를 찾을 수 없습니다: {match_value}\"
        )
    return [int(d) for d in digits[:num_questions]]


def update_excel_scores(
    excel_path: str | Path,
    id_to_score: dict[str, float],
    id_column: str = "사번",
    score_column: str = "점수",
    sheet_name: str | None = None,
    prefer_second_sheet: bool = False,
) -> None:
    """엑셀에서 id_column으로 사번을 찾아 해당 행의 score_column에 점수 기록."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {path}")
    wb = openpyxl.load_workbook(path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif prefer_second_sheet and len(wb.worksheets) >= 2:
        # 시트1 = 시험번호/정답, 시트2 = 응시생 명단이라는 가정일 때 사용
        ws = wb.worksheets[1]
    else:
        ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        id_col_idx = headers.index(id_column) + 1
    except ValueError:
        raise ValueError(f"엑셀 첫 행에 '{id_column}' 컬럼이 없습니다. 헤더: {headers}")

    if score_column not in headers:
        score_col_idx = len(headers) + 1
        ws.cell(row=1, column=score_col_idx, value=score_column)
    else:
        score_col_idx = headers.index(score_column) + 1

    for row_idx in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=id_col_idx).value
        if cell_id is None:
            continue
        sid = str(cell_id).strip()
        if sid in id_to_score:
            ws.cell(row=row_idx, column=score_col_idx, value=id_to_score[sid])

    wb.save(path)
    print(f"엑셀 저장 완료: {path}, 반영된 인원: {len(id_to_score)}명")


def try_decode_qr_from_image(image_path: str | Path) -> str | None:
    """이미지에서 QR코드 디코딩해 문자열 반환 (사번 등). 실패 시 None."""
    try:
        import cv2
        img = cv2.imread(str(image_path))
        if img is None:
            return None
        det = cv2.QRCodeDetector()
        data, _, _ = det.detectAndDecode(img)
        return data.strip() or None
    except Exception:
        return None


def main():
    import argparse
    parser = argparse.ArgumentParser(description="OMR 이미지 채점 후 엑셀에 점수 반영")
    parser.add_argument("images", nargs="+", help="OMR 카드 이미지 파일 경로들")
    parser.add_argument("--excel", "-e", required=True, help="응시생 정보 엑셀 파일 경로")
    parser.add_argument("--config", "-c", default="config.yaml", help="설정 파일 (기본: config.yaml)")
    parser.add_argument(
        "--exam-code",
        help="시험번호 (엑셀 시트1의 1열 값과 일치해야 함). 지정하면 시트1에서 시험정답을 자동으로 읽어 사용.",
    )
    parser.add_argument("--dry-run", action="store_true", help="엑셀 저장 없이 채점 결과만 출력")
    parser.add_argument("--use-qr", action="store_true", help="이미지에서 QR로 사번 읽기 시도")
    args = parser.parse_args()

    config = load_config(args.config)
    layout = config.get("layout", {})
    num_q = layout.get("num_questions", 40)
    choices = layout.get("choices_per_question", 5)
    scoring = config.get("scoring", {})
    points_per = scoring.get("points_per_question", 1.0)
    excel_cfg = config.get("excel", {})
    id_column = excel_cfg.get("id_column", "사번")
    score_column = excel_cfg.get("score_column", "점수")
    sheet_name = excel_cfg.get("sheet_name")

    # 정답 키 결정:
    # - --exam-code 가 지정되면: 엑셀 시트1에서 해당 시험번호의 정답 자동 읽기
    # - 아니면: config.yaml 의 answer_key 사용
    if args.exam_code:
        answer_key = load_answer_key_from_excel(args.excel, args.exam_code, num_q)
        print(f\"시험번호 {args.exam_code} 의 정답을 엑셀 시트1에서 읽었습니다.\")
    else:
        answer_key = config["answer_key"]

    id_to_score = {}
    for img_path in args.images:
        path = Path(img_path)
        if not path.exists():
            print(f"건너뜀 (파일 없음): {path}")
            continue
        # 사번 결정: QR 시도 후 파일명
        sid = None
        if args.use_qr:
            sid = try_decode_qr_from_image(path)
        if not sid:
            sid = get_id_from_filename(path.name)
        if not sid:
            print(f"건너뜀 (사번 추출 불가): {path}")
            continue
        answers = read_omr(path, num_questions=num_q, choices_per_question=choices)
        score = score_answers(answers, answer_key, points_per)
        id_to_score[sid] = score
        print(f"  {path.name} -> 사번 {sid}, 점수 {score:.1f}")

    if not id_to_score:
        print("채점된 답안이 없습니다.")
        sys.exit(1)

    if args.dry_run:
        print("--dry-run: 엑셀에 저장하지 않음.")
        return

    update_excel_scores(
        args.excel,
        id_to_score,
        id_column=id_column,
        score_column=score_column,
        sheet_name=sheet_name,
        prefer_second_sheet=bool(args.exam_code),
    )


if __name__ == "__main__":
    main()
